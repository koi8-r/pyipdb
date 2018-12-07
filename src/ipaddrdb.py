from openpyxl import load_workbook, Workbook
from itertools import count
import subprocess as sp
from ipaddress import IPv4Address


src = '../db/ipaddrdb.xlsx'
dst = '../db/ipaddrnew.xlsx'
check = 'PING'

sheet = load_workbook(src)['DB']
print('{}x{}'.format(sheet.max_column, sheet.max_row))

def _next(sheet):
    for row in count(start=2):
        line = tuple((sheet.cell(row=row, column=c).value or '').strip()
                     or None
                     for c in (1, 2, 3, ))
        if line == (None, None, None, ):
            break
        else:
            if check.upper() == 'PING':
                result = sp.run('ping -4 -n 1 -w 1500 ' + line[0],
                                shell=False,
                                stdout=sp.DEVNULL,
                                stderr=sp.DEVNULL)\
                            .returncode
            else:
                result = True
            yield line + (not result, )


wb = Workbook()
wb.create_sheet('DB')
for rec in sorted(_next(sheet), key=lambda t: IPv4Address(t[0])):
    print("'{}', '{}', '{}', {}".format(*rec))
    wb.active.append(rec)
wb.save(dst)
