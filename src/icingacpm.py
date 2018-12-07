from openpyxl import load_workbook, Workbook
from itertools import count
from ipaddress import IPv4Address


src = '../db/ip.xlsx'

db, icinga = [f[name] for f in [load_workbook(src)]
                      for name in ['DB', 'Icinga']]

def _next(sheet):
    for row in count(start=2):
        line = tuple((sheet.cell(row=row, column=c).value or '').strip().lower()
                     or None
                     for c in (1, 2, 3, ))
        if line == (None, None, None, ):
            break
        else:
            yield line


db = [rec[0] for rec in _next(db)]
for rec in sorted(_next(icinga), key=lambda t: IPv4Address(t[0])):
    if rec[0] not in db:
        print("{}, {}, {}".format(*rec))
