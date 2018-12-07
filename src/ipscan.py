from openpyxl import load_workbook, Workbook
from itertools import count
import subprocess as sp
from ipaddress import IPv4Address, IPv4Network
import socket


db = '../db/ip.xlsx'
sheet = load_workbook(db)['DB']

def _next(sheet):
    for row in count(start=2):
        line = tuple((sheet.cell(row=row, column=c).value or '').strip()
                     or None
                     for c in (1, 2, 3, ))
        if line == (None, None, None, ):
            break
        else:
            yield IPv4Address(line[0])


def check_port(addr, port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(0.5)
    try:
        sock.connect((socket.gethostbyname(addr), port, ))
    except:
        return None
    finally:
        sock.settimeout(None)
        sock.close()    
    return port


learned =  list(_next(sheet))
for ip in (ip for net in [
    '10.64.20.0/24',
    '10.64.21.0/24',
    '10.64.23.0/24',
    '10.64.32.0/24',
    '10.64.136.0/24',
    '10.64.137.0/24',
    '10.64.138.0/24',
    '10.64.139.0/24',
    '10.64.140.0/24',
    '10.64.141.0/24',
    '10.64.142.0/24',
    '10.64.143.0/24',
    '172.18.28.0/24',
    '172.18.29.0/24',
    '172.18.96.0/24',
    '172.18.97.0/24',
    '172.18.98.0/24',
    '172.18.99.0/24',
    '172.18.232.0/24',
    '172.18.242.0/24',
    '172.24.86.0/24',
    '172.26.87.0/24',
    '172.27.5.0/24',
    '172.31.21.0/24',
] for ip in IPv4Network(net)):
    if ip not in learned \
       and not sp.run('ping -4 -n 1 -w 1500 ' + str(ip),
                      shell=False,
                      stdout=sp.DEVNULL,
                      stderr=sp.DEVNULL)\
                      .returncode:
        print((str(ip) + ' ' + ' '.join(str(p) for p in [80, 22, 443] if check_port(str(ip), p))).rstrip())
