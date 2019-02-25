from openpyxl import load_workbook, Workbook
from itertools import count
from ipaddress import IPv4Address


src = '../db/ipaddrdb.xlsx'
wb = load_workbook(src)
db, dbnew = [f[name] for f in [wb]
                     for name in ['DB', 'DBNEW']]

def _next(sheet):
    for row in count(start=2):
        line = tuple(sheet.cell(row=row, column=c).value
                     for c in range(1, 5+1))
        if line == tuple(None for _ in range(5)):
            break
        else:
            yield line


for rec in sorted(_next(db), key=lambda t: IPv4Address(t[0])):
    print("{}, {}, {}, {}, {}".format(*rec))
    dbnew.append(rec)
wb.save(src)
